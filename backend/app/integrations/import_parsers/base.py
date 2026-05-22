import csv
from abc import ABC, abstractmethod
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from typing import Any
from uuid import UUID

from app.domain.enums import (
    ImportFileType,
    ImportSourceType,
    TransactionDirection,
    TransactionType,
)
from app.domain.ledger import (
    ImportFileMetadata,
    ImportValidationError,
    NormalizedTransactionCandidate,
    ParsedRawTransaction,
)


@dataclass(frozen=True)
class ImportParseContext:
    import_batch_id: UUID
    user_id: UUID
    account_id: UUID
    source_type: ImportSourceType
    provider: str | None = None


class FinancialImportParser(ABC):
    provider_id: str
    supported_file_types: set[ImportFileType]

    def can_parse(self, file_metadata: ImportFileMetadata, provider: str | None) -> bool:
        provider_match = provider in {None, "", self.provider_id, "generic"}
        return file_metadata.file_type in self.supported_file_types and provider_match

    @abstractmethod
    def parse(
        self,
        file_content: bytes,
        context: ImportParseContext,
        file_metadata: ImportFileMetadata,
    ) -> list[ParsedRawTransaction]:
        ...

    @abstractmethod
    def validate_row(self, row: ParsedRawTransaction) -> list[ImportValidationError]:
        ...

    @abstractmethod
    def normalize_row(
        self,
        row: ParsedRawTransaction,
        context: ImportParseContext,
    ) -> NormalizedTransactionCandidate:
        ...


class TabularImportParser(FinancialImportParser):
    def parse(
        self,
        file_content: bytes,
        context: ImportParseContext,
        file_metadata: ImportFileMetadata,
    ) -> list[ParsedRawTransaction]:
        if file_metadata.file_type == ImportFileType.csv:
            rows = self._read_csv(file_content)
        elif file_metadata.file_type == ImportFileType.xlsx:
            rows = self._read_xlsx(file_content)
        else:
            raise ValueError("XLS imports need the future binary spreadsheet adapter.")

        return [
            ParsedRawTransaction(row_number=index + 2, raw=row)
            for index, row in enumerate(rows)
            if any(str(value).strip() for value in row.values() if value is not None)
        ]

    def _read_csv(self, file_content: bytes) -> list[dict[str, Any]]:
        text = file_content.decode("utf-8-sig")
        return [dict(row) for row in csv.DictReader(StringIO(text))]

    def _read_xlsx(self, file_content: bytes) -> list[dict[str, Any]]:
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(file_content), data_only=True, read_only=True)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        parsed_rows = []
        for row in rows[1:]:
            parsed_rows.append(
                {
                    header: value
                    for header, value in zip(headers, row, strict=False)
                    if header
                }
            )

        return parsed_rows

    def _value(self, raw: dict[str, Any], names: list[str]) -> Any:
        normalized = {self._normalize_key(key): value for key, value in raw.items()}
        for name in names:
            value = normalized.get(self._normalize_key(name))
            if value not in {None, ""}:
                return value
        return None

    def _required_errors(self, row: ParsedRawTransaction, fields: dict[str, Any]) -> list[ImportValidationError]:
        errors = []
        for field_name, value in fields.items():
            if value in {None, ""}:
                errors.append(
                    ImportValidationError(
                        row_number=row.row_number,
                        field=field_name,
                        message=f"{field_name} is required.",
                    )
                )
        return errors

    def _parse_date(self, value: Any) -> date:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value

        text = str(value).strip()
        for date_format in ("%Y-%m-%d", "%d/%m/%Y", "%d.%m.%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(text, date_format).date()
            except ValueError:
                continue

        raise ValueError("transaction_date must be a valid date.")

    def _parse_amount(self, value: Any) -> Decimal:
        try:
            amount = Decimal(str(value).replace(",", "").strip())
        except (InvalidOperation, ValueError) as error:
            raise ValueError("amount must be numeric.") from error

        if amount == 0:
            raise ValueError("transaction amount cannot be zero.")

        return abs(amount)

    def _direction_from_amount(self, amount: Any, default: TransactionDirection) -> TransactionDirection:
        try:
            parsed = Decimal(str(amount).replace(",", "").strip())
        except (InvalidOperation, ValueError):
            return default
        return TransactionDirection.credit if parsed > 0 else TransactionDirection.debit

    def _normalize_key(self, key: str) -> str:
        return "".join(str(key).lower().split())

    def _normalized_description(self, value: Any) -> str:
        return " ".join(str(value or "").strip().split())

    def _transaction_type_from_direction(
        self,
        direction: TransactionDirection,
    ) -> TransactionType:
        return TransactionType.income if direction == TransactionDirection.credit else TransactionType.expense
